"""
PDF → Excel (.xlsx)
Uses camelot to detect tables in PDF and write each table to a separate sheet.
Works best on PDFs with clear table borders. Falls back to plain text extraction.
"""

import openpyxl
from openpyxl.utils.dataframe import dataframe_to_rows


def convert(input_path, output_path, options):
    output = output_path + ".xlsx"

    try:
        import camelot
        tables = camelot.read_pdf(input_path, pages="all", flavor="lattice")

        if len(tables) == 0:
            # Try stream flavor (for borderless tables)
            tables = camelot.read_pdf(input_path, pages="all", flavor="stream")
    except Exception:
        tables = []

    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    if len(tables) == 0:
        # No tables found - just dump text per page
        from pypdf import PdfReader
        reader = PdfReader(input_path)
        ws = wb.create_sheet("Extracted Text")
        for i, page in enumerate(reader.pages):
            ws.append([f"--- Page {i+1} ---"])
            text = page.extract_text() or ""
            for line in text.split("\n"):
                ws.append([line])
    else:
        for i, table in enumerate(tables):
            ws = wb.create_sheet(f"Table_{i+1}")
            df = table.df
            for row in dataframe_to_rows(df, index=False, header=False):
                ws.append(row)

    wb.save(output)
    return output
